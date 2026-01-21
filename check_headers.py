import openpyxl

wb = openpyxl.load_workbook('MATRIZ DE LIMITACIONES LABORALES.xlsx')
ws = wb['TEMPORAL Y DEFINITIVA']

with open('all_headers.txt', 'w', encoding='utf-8') as f:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        # Si fila 2 está vacía, intentar fila 1 o 3
        if not val:
            val = ws.cell(row=1, column=col).value
        
        f.write(f"Col {col}: {val}\n")
        
        # Muestreo de datos de esa columna (filas 3-5)
        for r in range(3, 6):
            cell_val = ws.cell(row=r, column=col).value
            if cell_val:
                f.write(f"   Row {r}: {cell_val}\n")

print("Headers saved to all_headers.txt")
