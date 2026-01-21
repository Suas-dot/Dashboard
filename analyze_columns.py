import openpyxl

wb = openpyxl.load_workbook('MATRIZ DE LIMITACIONES LABORALES.xlsx', data_only=True)
ws = wb['TEMPORAL Y DEFINITIVA']

print("ANALIZANDO COLUMNAS ESPECÍFICAS DE RESTRICCIONES:")
print("-" * 50)

# Columnas identificadas (Excel index based)
# Col 33: No levantar cargar > a 5 kg.
# Col 34: No levantar cargar > a 15 kg.
# Col 47: No realizar trabajos nocturnos 
# Col 57: Desvinculación

col_peso_5 = 33
col_peso_15 = 34
col_nocturno = 47
col_desvinculacion = 57

count_peso_5 = 0
count_peso_15 = 0
count_nocturno = 0
count_desvinculacion = 0

desvinculacion_details = []

for row in range(3, ws.max_row + 1):
    # Verificar si es un empleado válido (tiene código)
    codigo = ws.cell(row=row, column=2).value
    if not codigo: continue
    
    # Peso > 5kg
    val_5 = ws.cell(row=row, column=col_peso_5).value
    if val_5 and str(val_5).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_peso_5 += 1
        
    # Peso > 15kg
    val_15 = ws.cell(row=row, column=col_peso_15).value
    if val_15 and str(val_15).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_peso_15 += 1
        
    # Nocturno
    val_noc = ws.cell(row=row, column=col_nocturno).value
    if val_noc and str(val_noc).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_nocturno += 1
        
    # Desvinculación
    val_des = ws.cell(row=row, column=col_desvinculacion).value
    if val_des and str(val_des).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_desvinculacion += 1
        nombre = ws.cell(row=row, column=4).value
        apellido = ws.cell(row=row, column=5).value
        desvinculacion_details.append(f"{nombre} {apellido} ({codigo})")

print(f"🏋️  Restricción Carga > 5 kg: {count_peso_5}")
print(f"🏋️  Restricción Carga > 15 kg: {count_peso_15}")
print(f"    -> Total restricción Cargas: {count_peso_5 + count_peso_15} (Puede haber solapamiento)")
print("-" * 30)
print(f"🌙  Restricción Turnos Nocturnos: {count_nocturno}")
print("-" * 30)
print(f"🚫  Desvinculación: {count_desvinculacion}")

if count_desvinculacion > 0:
    print("\nDetalle de posibles desvinculaciones:")
    for p in desvinculacion_details:
        print(f"  - {p}")
