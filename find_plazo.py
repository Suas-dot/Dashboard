import openpyxl

wb = openpyxl.load_workbook('MATRIZ DE LIMITACIONES LABORALES.xlsx', data_only=True)

print("BUSCANDO 'PLAZO' EN TODO EL ARCHIVO:")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n--- {sheet} ---")
    for r in range(1, min(ws.max_row+1, 100)):
        for c in range(1, min(ws.max_column+1, 100)):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and 'PLAZO' in val.upper():
                print(f"ENCONTRADO en {sheet} [{r},{c}]: {val}")
