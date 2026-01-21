import openpyxl

wb = openpyxl.load_workbook('MATRIZ DE LIMITACIONES LABORALES.xlsx', data_only=True)
ws = wb['TEMPORAL Y DEFINITIVA']

col_peso_5 = 33
col_peso_15 = 34
col_nocturno = 47
col_desvinculacion = 57

count_peso_5 = 0
count_peso_15 = 0
count_nocturno = 0
count_desvinculacion = 0

for row in range(3, ws.max_row + 1):
    val_5 = ws.cell(row=row, column=col_peso_5).value
    if val_5 and str(val_5).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_peso_5 += 1
        
    val_15 = ws.cell(row=row, column=col_peso_15).value
    if val_15 and str(val_15).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_peso_15 += 1
        
    val_noc = ws.cell(row=row, column=col_nocturno).value
    if val_noc and str(val_noc).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_nocturno += 1
        
    val_des = ws.cell(row=row, column=col_desvinculacion).value
    if val_des and str(val_des).strip().upper() in ['TRUE', 'SI', 'X', '1']:
        count_desvinculacion += 1

print("\n\n" + "="*50)
print(" RESULTADOS DEL ANÁLISIS")
print("="*50)
print(f"🚫  DESVINCULACIÓN: {count_desvinculacion}")
print(f"🏋️  RESTRICCIÓN CARGA > 5 kg: {count_peso_5}")
print(f"🏋️  RESTRICCIÓN CARGA > 15 kg: {count_peso_15}")
print(f"🌙  RESTRICCIÓN TURNOS NOCTURNOS: {count_nocturno}")
print("="*50 + "\n\n")
